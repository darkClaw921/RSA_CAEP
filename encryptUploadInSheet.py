import gspread
from cryptoMessage import encrypt_message

from openpyxl import load_workbook
from oauth2client.service_account import ServiceAccountCredentials

print('Создание таблицы ...')
SCOPE = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive'] # что то для чего-то нужно Костыль    
CREDS = ServiceAccountCredentials.from_json_keyfile_name("/Users/igorgerasimov/Desktop/Мусор/", SCOPE) # Секретынй файл json для доступа к API
CLIENT = gspread.authorize(CREDS)
SHEET = CLIENT.open('CAEP_KGTA').sheet1 # Имя таблицы
print('Таблица создана')

WB = load_workbook('/Users/igorgerasimov/Desktop/test.xlsx')
SHEET_DATE_BASE = WB.active

SHEET.update_cell(1, 1, 'ФИО')
SHEET.update_cell(1, 2, 'Логин')
SHEET.update_cell(1, 3, 'Пароль')

rowSheetDataBase = 2
rowSheetGoogle = 2

for _ in range(16): #217

    print(f'{_} iteration')
    name = SHEET_DATE_BASE.cell(row=rowSheetDataBase, column=3).value
    login = SHEET_DATE_BASE.cell(row=rowSheetDataBase, column=6).value
    password = SHEET_DATE_BASE.cell(row=rowSheetDataBase, column=7).value

    if login is None or login == 'Нету ':
        rowSheetDataBase +=1
        continue

    nameEncrypt = encrypt_message(name)
    loginEncrypt = encrypt_message(login)
    passwordEncrypt = encrypt_message(password)
    
    SHEET.update_cell(rowSheetGoogle, 1, name)
    SHEET.update_cell(rowSheetGoogle, 2, str(loginEncrypt))
    SHEET.update_cell(rowSheetGoogle, 3, str(passwordEncrypt))

    rowSheetDataBase +=1
    rowSheetGoogle +=1
