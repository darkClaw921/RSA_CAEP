import gspread

from Crypto.PublicKey import RSA
from openpyxl import load_workbook, Workbook
from oauth2client.service_account import ServiceAccountCredentials

from cryptoMessage import decrypt_message

print('Подключение к таблице')
SCOPE = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive'] # что то для чего-то нужно Костыль    
CREDS = ServiceAccountCredentials.from_json_keyfile_name("", SCOPE) # Секретынй файл json для доступа к API
CLIENT = gspread.authorize(CREDS)
SHEET = CLIENT.open('CAEP_KGTA').sheet1 # Имя таблицы
print('[OK]')

wb = Workbook()

nameWorkBook = 'encryptSheet.xlsx'
DECRYPT_SHEET = wb.create_sheet()
DECRYPT_SHEET.title = 'date users'

name = SHEET.get('A1')
row = 2

while name[0][0] is not None:
    try:
        name = SHEET.get(f'A{row}')
        login = SHEET.get(f'B{row}')
        password = SHEET.get(f'C{row}')
    except KeyError: 
        print('Конец таблицы')
        break

    loginDecrypt = decrypt_message(login[0][0])
    passwordDecrypt = decrypt_message(password[0][0])

    DECRYPT_SHEET[f'A{row}'] = name[0][0]
    DECRYPT_SHEET[f'B{row}'] = loginDecrypt
    DECRYPT_SHEET[f'C{row}'] = passwordDecrypt

    row +=1
wb.save(nameWorkBook)
print('[+]   Декодирование завершино')








